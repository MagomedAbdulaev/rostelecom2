from django.shortcuts import render, redirect
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from .serializers import UserSerializer
from django.contrib.auth.decorators import login_required
from .utils import jwt_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate
from rest_framework_simplejwt.tokens import RefreshToken
from django.http import HttpResponseRedirect, JsonResponse
from .models import *
from django.contrib import messages
from decimal import Decimal
import openpyxl
from django.http import HttpResponse
from .forms import ExcelImportForm
from openpyxl.utils import get_column_letter
from django.contrib.auth import get_user_model
from django.db.models import Avg
from django.utils import timezone
from datetime import timedelta


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    """Возвращает информацию о текущем пользователе"""
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


def logout_user(request):
    response = redirect('rost:login_user')
    response.delete_cookie('access_token')
    response.delete_cookie('refresh_token')
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    """Возвращает информацию о текущем пользователе"""
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


def logout_user(request):
    response = redirect('rost:login_user')
    response.delete_cookie('access_token')
    response.delete_cookie('refresh_token')
    return response

@jwt_required
def home(request):
    """Главная страница - доступна всем авторизованным пользователям"""
    user = request.user
    
    # Получаем параметры фильтрации из GET запроса
    status_filter = request.GET.get('status', '')
    city_filter = request.GET.get('city', '')
    service_filter = request.GET.get('service', '')
    
    # Базовый queryset
    objects = Object.objects.select_related('city').order_by('-created_at')
    
    # Применяем фильтры
    if status_filter:
        objects = objects.filter(status=status_filter)
    if city_filter:
        objects = objects.filter(city_id=city_filter)
    if service_filter:
        # Исправляем проблему с JSONField
        objects = objects.filter(interested_services__isnull=False)
        objects = [obj for obj in objects if obj.interested_services and service_filter in obj.interested_services]
    
    # Ограничиваем количество для главной страницы, если нет фильтров
    if not any([status_filter, city_filter, service_filter]):
        objects = objects[:10]
    
    # Статистика (всегда считаем по всем объектам)
    objects_count = Object.objects.count()
    new_clients_count = Object.objects.filter(status='new').count()
    callback_count = Object.objects.filter(status='callback').count()
    potential_count = Object.objects.filter(status='potential').count()
    
    # Данные для фильтров
    cities = City.objects.all()
    
    context = {
        'objects': objects,
        'objects_count': objects_count,
        'new_clients_count': new_clients_count,
        'callback_count': callback_count,
        'potential_count': potential_count,
        'cities': cities,
        'current_status': status_filter,
        'current_city': city_filter,
        'current_service': service_filter,
        'title': 'Главная',
        'is_admin': hasattr(user, 'role') and user.role in ['admin', 'supervisor']  # Добавляем флаг для шаблона
    }
    return render(request, 'rost/home.html', context)


def import_admin_excel(request):
    if request.method == "POST":
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]

            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                headers = [str(h).strip() for h in rows[0]]
                User = get_user_model()

                for row in rows[1:]:
                    data = dict(zip(headers, row))

                    # Город
                    city_name = data.get("Город") or "Неизвестно"
                    city, _ = City.objects.get_or_create(name=city_name)

                    # Создатель
                    created_by_name = data.get("Кем создано")
                    created_by = None
                    if created_by_name:
                        created_by = User.objects.filter(username=created_by_name).first()

                    # Преобразование полей
                    services_used = (
                        [s.strip() for s in str(data.get("Используемые услуги", "")).split(",") if s.strip()]
                        if data.get("Используемые услуги") else None
                    )
                    interested_services = (
                        [s.strip() for s in str(data.get("Интересующие услуги", "")).split(",") if s.strip()]
                        if data.get("Интересующие услуги") else None
                    )

                    # Обновление / создание объекта
                    Object.objects.update_or_create(
                        id=data.get("ID"),
                        defaults={
                            "name": data.get("Название объекта"),
                            "city": city,
                            "address": data.get("Адрес"),
                            "apartment_number": data.get("Номер квартиры"),
                            "contact_phone": data.get("Телефон"),
                            "client_portrait": data.get("Портрет клиента"),
                            "services_used": services_used,
                            "interested_services": interested_services,
                            "current_provider_rating": data.get("Оценка провайдера") or None,
                            "convenient_contact_time": None,  # можно доработать парсер
                            "desired_price": data.get("Желаемая цена") or None,
                            "notes": data.get("Примечание"),
                            "latitude": data.get("Широта") or None,
                            "longitude": data.get("Долгота") or None,
                            "created_by": created_by,
                            "status": (data.get("Статус") or "new").lower(),
                        },
                    )

                messages.success(request, "✅ Импорт успешно завершён.")
            except Exception as e:
                messages.error(request, f"❌ Ошибка при импорте: {e}")

            return redirect("rost:admin_home")
    else:
        form = ExcelImportForm()

    return render(request, "rost/import_excel.html", {"form": form})


def export_admin_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Объекты"

    # Заголовки
    headers = [
        "ID", "Название объекта", "Город", "Адрес", "Номер квартиры",
        "Телефон", "Портрет клиента", "Используемые услуги", "Интересующие услуги",
        "Оценка провайдера", "Удобное время", "Желаемая цена", "Примечание",
        "Широта", "Долгота", "Кем создано", "Статус", "Дата следующего контакта",
        "Дата создания", "Дата обновления"
    ]
    ws.append(headers)

    # Данные
    for obj in Object.objects.select_related("city", "created_by"):
        ws.append([
            obj.id,
            obj.name,
            obj.city.name if obj.city else "",
            obj.address,
            obj.apartment_number,
            obj.contact_phone,
            obj.client_portrait,
            ", ".join(obj.services_used or []) if obj.services_used else "",
            ", ".join(obj.interested_services or []) if obj.interested_services else "",
            obj.current_provider_rating or "",
            obj.get_convenient_contact_time_display() if obj.convenient_contact_time else "",
            obj.desired_price or "",
            obj.notes,
            obj.latitude or "",
            obj.longitude or "",
            obj.created_by.username if obj.created_by else "",
            obj.status,
            obj.next_contact_date.strftime("%d.%m.%Y") if obj.next_contact_date else "",
            obj.created_at.strftime("%d.%m.%Y %H:%M"),
            obj.updated_at.strftime("%d.%m.%Y %H:%M"),
        ])

    # Красота: автоширина столбцов
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=objects_export.xlsx'
    wb.save(response)
    return response


@jwt_required
def admin_home(request):
    user = request.user
    if not hasattr(user, 'role') or user.role not in ['admin', 'supervisor']:
        messages.error(request, 'У вас нет доступа к админ панели')
        return redirect('rost:home')

    # Основная статистика
    objects_count = Object.objects.count()
    new_clients_count = Object.objects.filter(status='new').count()
    callback_count = Object.objects.filter(status='callback').count()
    potential_count = Object.objects.filter(status='potential').count()
    refused_count = Object.objects.filter(status='refused').count()
    existing_count = Object.objects.filter(status='existing').count()

    # Статистика по городам
    cities_stats = City.objects.annotate(objects_count=models.Count('building_objects')).order_by('-objects_count')

    # Последние объекты
    recent_objects = Object.objects.select_related('city', 'created_by').order_by('-created_at')[:10]

    # Статистика по услугам
    services_stats_display = {}
    all_objects = Object.objects.all()
    for service_code, service_name in Object.INTEREST_SERVICES_CHOICES:
        count = sum(1 for obj in all_objects if obj.interested_services and service_code in obj.interested_services)
        services_stats_display[service_name] = count

    # --- Удовлетворенность клиентов ---
    period = request.GET.get('satisfaction_period', 'month')  # день/неделя/месяц/год
    now = timezone.now()
    if period == 'day':
        start_date = now - timedelta(days=1)
    elif period == 'week':
        start_date = now - timedelta(weeks=1)
    elif period == 'month':
        start_date = now - timedelta(days=30)
    elif period == 'year':
        start_date = now - timedelta(days=365)
    else:
        start_date = now - timedelta(days=30)

    satisfaction_qs = Object.objects.filter(
        created_at__gte=start_date,
        current_provider_rating__isnull=False
    ).order_by('created_at')

    satisfaction_labels = [obj.created_at.strftime('%d.%m') for obj in satisfaction_qs]
    satisfaction_values = [obj.current_provider_rating for obj in satisfaction_qs]

    context = {
        'objects_count': objects_count,
        'new_clients_count': new_clients_count,
        'callback_count': callback_count,
        'potential_count': potential_count,
        'refused_count': refused_count,
        'existing_count': existing_count,
        'cities_stats': cities_stats,
        'recent_objects': recent_objects,
        'services_stats': services_stats_display,
        'title': 'Админ панель',
        'satisfaction_chart_labels': satisfaction_labels,
        'satisfaction_chart_values': satisfaction_values,
        'satisfaction_period': period
    }
    return render(request, 'rost/admin_home.html', context)
    

@jwt_required
def create_object(request):
    if request.method == 'POST':
        try:
            city = City.objects.get(id=request.POST.get('city'))
            name = request.POST.get('name')
            address = request.POST.get('address')
            latitude = request.POST.get('latitude') or None
            longitude = request.POST.get('longitude') or None
            
            # Получаем списки из чекбоксов
            services_used = request.POST.getlist('services_used')
            interested_services = request.POST.getlist('interested_services')
            
            # Получаем числовое значение для рейтинга
            current_provider_rating = request.POST.get('current_provider_rating')
            if current_provider_rating:
                current_provider_rating = int(current_provider_rating)
            
            # Получаем decimal для цены
            desired_price = request.POST.get('desired_price')
            if desired_price:
                desired_price = Decimal(desired_price)
            
            obj = Object.objects.create(
                city=city,
                name=name,
                address=address,
                apartment_number=request.POST.get('apartment_number') or None,
                contact_phone=request.POST.get('contact_phone') or None,
                client_portrait=request.POST.get('client_portrait') or None,
                services_used=services_used,  # уже список
                current_provider_rating=current_provider_rating,
                interested_services=interested_services,  # уже список
                convenient_contact_time=request.POST.get('convenient_contact_time') or None,
                desired_price=desired_price,
                notes=request.POST.get('notes') or None,
                latitude=latitude,
                longitude=longitude,
                status=request.POST.get('status', 'new'),  # добавляем статус
                created_by=request.user
            )
            
            messages.success(request, f'Объект "{obj.name}" успешно создан!')
            return redirect('rost:home')
            
        except Exception as e:
            pass
    
    cities = City.objects.all()
    satisfaction_choices = [
        (1, '1 - Очень неудовлетворен'),
        (2, '2 - Неудовлетворен'),
        (3, '3 - Нейтрален'),
        (4, '4 - Удовлетворен'),
        (5, '5 - Очень удовлетворен'),
    ]
    context = {
        'cities': cities,
        'satisfaction_choices': satisfaction_choices,
        'title': 'Добавление объекта'
    }
    return render(request, 'rost/create_object.html', context=context)


@jwt_required
def edit_object(request, object_id):
    try:
        obj = Object.objects.get(id=object_id)
    except Object.DoesNotExist:
        messages.error(request, 'Объект не найден')
        return redirect('rost:home')

    if request.method == 'POST':
        try:
            city = City.objects.get(id=request.POST.get('city'))
            
            # Получаем списки из чекбоксов
            services_used = request.POST.getlist('services_used')
            interested_services = request.POST.getlist('interested_services')
            
            # Получаем числовое значение для рейтинга
            current_provider_rating = request.POST.get('current_provider_rating')
            if current_provider_rating:
                current_provider_rating = int(current_provider_rating)
            else:
                current_provider_rating = None
            
            # Получаем decimal для цены - исправленная обработка
            desired_price = request.POST.get('desired_price')
            if desired_price and desired_price.strip():  # Проверяем, что не пустая строка
                desired_price = Decimal(desired_price)
            else:
                desired_price = None
            
            # Обновляем объект
            obj.city = city
            obj.name = request.POST.get('name')
            obj.address = request.POST.get('address')
            obj.apartment_number = request.POST.get('apartment_number') or None
            obj.contact_phone = request.POST.get('contact_phone') or None
            obj.client_portrait = request.POST.get('client_portrait') or None
            obj.services_used = services_used
            obj.current_provider_rating = current_provider_rating
            obj.interested_services = interested_services
            obj.convenient_contact_time = request.POST.get('convenient_contact_time') or None
            obj.desired_price = desired_price  # Может быть None
            obj.notes = request.POST.get('notes') or None
            obj.latitude = request.POST.get('latitude') or None
            obj.longitude = request.POST.get('longitude') or None
            obj.status = request.POST.get('status', 'new')
            
            obj.save()
            
            messages.success(request, f'Объект "{obj.name}" успешно обновлен!')
            return redirect('rost:home')
            
        except Exception as e:
            print(f"Ошибка при обновлении объекта: {str(e)}")
            messages.error(request, f'Ошибка при обновлении объекта: {str(e)}')

    cities = City.objects.all()
    satisfaction_choices = [
        (1, '1 - Очень неудовлетворен'),
        (2, '2 - Неудовлетворен'),
        (3, '3 - Нейтрален'),
        (4, '4 - Удовлетворен'),
        (5, '5 - Очень удовлетворен'),
    ]
    
    context = {
        'object': obj,
        'price': str(obj.desired_price).split('.')[0],
        'cities': cities,
        'satisfaction_choices': satisfaction_choices,
        'title': f'Редактирование {obj.name}'
    }
    return render(request, 'rost/edit_object.html', context)


@jwt_required
def delete_object(request, object_id):
    if request.method == 'DELETE':
        try:
            obj = Object.objects.get(id=object_id)
            obj_name = obj.name
            obj.delete()
            return JsonResponse({'success': True, 'message': f'Объект "{obj_name}" удален'})
        except Object.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Объект не найден'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Метод не разрешен'}, status=405)


@csrf_exempt
def login_user(request):
    """JWT логин — выдаёт токен при правильных данных"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)

            # сохраняем токен в cookie
            response = redirect('rost:home')
            response.set_cookie(
                key='access_token',
                value=access_token,
                httponly=True,
                secure=False,   # True если HTTPS
                samesite='Lax'
            )
            response.set_cookie(
                key='refresh_token',
                value=str(refresh),
                httponly=True,
                secure=False,
                samesite='Lax'
            )
            return response

        # ошибка входа
        return render(request, 'rost/login.html', {
            'title': 'Авторизация',
            'form': {'errors': True}
        })

    return render(request, 'rost/login.html', {'title': 'Авторизация'})
